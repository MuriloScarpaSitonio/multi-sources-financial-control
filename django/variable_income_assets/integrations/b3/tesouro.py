from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .parser import B3ParserError
from .schemas import B3FixedIncomeAction, B3TesouroMovement, B3TesouroPosition

POSICAO_SHEET = "Tesouro Direto"
MOVIMENTACAO_SHEET = "Movimentação"
TESOURO_PRODUTO_PREFIX = "Tesouro"
COMPRA_VENDA_LABELS = {
    "Compra": B3FixedIncomeAction.BUY,
    "Venda": B3FixedIncomeAction.SELL,
}
ACTION_BY_FLOW = {"Credito": B3FixedIncomeAction.BUY, "Debito": B3FixedIncomeAction.SELL}

POSICAO_REQUIRED_HEADERS = (
    "Produto",
    "Código ISIN",
    "Indexador",
    "Vencimento",
    "Quantidade",
    "Valor Atualizado",
)

MOVIMENTACAO_REQUIRED_HEADERS = (
    "Entrada/Saída",
    "Data",
    "Movimentação",
    "Produto",
    "Quantidade",
    "Preço unitário",
)


def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() in ("", "-"):
        return True
    return False


def _to_optional_str(value) -> str | None:
    if _is_blank(value):
        return None
    return str(value).strip()


def _to_required_str(value, *, column: str, row_index: int) -> str:
    if _is_blank(value):
        raise B3ParserError(f"row {row_index}: required column {column!r} is blank")
    return str(value).strip()


def _to_optional_date(value, *, column: str, row_index: int) -> date | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def _to_required_date(value, *, column: str, row_index: int) -> date:
    if _is_blank(value):
        raise B3ParserError(f"row {row_index}: required column {column!r} is blank")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
    except ValueError as exc:
        raise B3ParserError(
            f"row {row_index}: could not parse date in {column!r}: {value!r}"
        ) from exc


def _to_required_decimal(value, *, column: str, row_index: int) -> Decimal:
    if _is_blank(value):
        raise B3ParserError(f"row {row_index}: required column {column!r} is blank")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise B3ParserError(
            f"row {row_index}: could not parse decimal in {column!r}: {value!r}"
        ) from exc


def _to_optional_decimal(value, *, column: str, row_index: int) -> Decimal | None:
    if _is_blank(value):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _build_header_index(header_row: tuple, *, required: tuple[str, ...]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        index[str(cell).strip()] = i

    for header in required:
        if header not in index:
            raise B3ParserError(f"missing column: {header}")

    return index


def parse_tesouro_positions(path: str) -> list[B3TesouroPosition]:
    resolved = Path(path)
    workbook = load_workbook(resolved, data_only=True)
    try:
        if POSICAO_SHEET not in workbook.sheetnames:
            raise B3ParserError(f"sheet {POSICAO_SHEET!r} not found in {resolved}")

        sheet = workbook[POSICAO_SHEET]
        rows = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise B3ParserError(f"empty sheet {POSICAO_SHEET!r}") from exc

        h = _build_header_index(header_row, required=POSICAO_REQUIRED_HEADERS)
        positions: list[B3TesouroPosition] = []

        for row_index, row in enumerate(rows, start=2):
            produto = row[h["Produto"]] if h["Produto"] < len(row) else None
            if _is_blank(produto):
                continue

            name = _to_required_str(produto, column="Produto", row_index=row_index)
            positions.append(
                B3TesouroPosition(
                    name=name,
                    isin=_to_required_str(
                        row[h["Código ISIN"]], column="Código ISIN", row_index=row_index
                    ),
                    indexer=_to_optional_str(row[h["Indexador"]]),
                    maturity_date=_to_optional_date(
                        row[h["Vencimento"]], column="Vencimento", row_index=row_index
                    ),
                    quantity=_to_required_decimal(
                        row[h["Quantidade"]], column="Quantidade", row_index=row_index
                    ),
                    current_value=_to_optional_decimal(
                        row[h["Valor Atualizado"]],
                        column="Valor Atualizado",
                        row_index=row_index,
                    ),
                )
            )

        return positions
    finally:
        workbook.close()


def parse_tesouro_movements(path: str) -> list[B3TesouroMovement]:
    resolved = Path(path)
    workbook = load_workbook(resolved, data_only=True)
    try:
        if MOVIMENTACAO_SHEET not in workbook.sheetnames:
            raise B3ParserError(f"sheet {MOVIMENTACAO_SHEET!r} not found in {resolved}")

        sheet = workbook[MOVIMENTACAO_SHEET]
        rows = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise B3ParserError(f"empty sheet {MOVIMENTACAO_SHEET!r}") from exc

        h = _build_header_index(header_row, required=MOVIMENTACAO_REQUIRED_HEADERS)
        movements: list[B3TesouroMovement] = []

        for row_index, row in enumerate(rows, start=2):
            produto_raw = row[h["Produto"]] if h["Produto"] < len(row) else None
            if _is_blank(produto_raw):
                continue

            produto = str(produto_raw).strip()
            if not produto.startswith(TESOURO_PRODUTO_PREFIX):
                continue

            label = str(row[h["Movimentação"]] or "").strip()
            action = COMPRA_VENDA_LABELS.get(label)
            if action is None:
                continue

            flow = _to_required_str(
                row[h["Entrada/Saída"]], column="Entrada/Saída", row_index=row_index
            )
            if ACTION_BY_FLOW.get(flow) != action:
                raise B3ParserError(
                    f"row {row_index}: flow {flow!r} disagrees with movimentação {label!r}"
                )

            movements.append(
                B3TesouroMovement(
                    name=produto,
                    action=action,
                    operation_date=_to_required_date(
                        row[h["Data"]], column="Data", row_index=row_index
                    ),
                    quantity=_to_required_decimal(
                        row[h["Quantidade"]], column="Quantidade", row_index=row_index
                    ),
                    unit_price=_to_required_decimal(
                        row[h["Preço unitário"]], column="Preço unitário", row_index=row_index
                    ),
                )
            )

        return movements
    finally:
        workbook.close()
