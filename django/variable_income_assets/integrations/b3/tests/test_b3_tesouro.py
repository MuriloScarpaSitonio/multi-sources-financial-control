from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from ..parser import B3ParserError
from ..schemas import B3FixedIncomeAction
from ..tesouro import parse_tesouro_movements, parse_tesouro_positions

POSICAO_HEADER = [
    "Produto",
    "Instituição",
    "Código ISIN",
    "Indexador",
    "Vencimento",
    "Quantidade",
    "Quantidade Disponível",
    "Quantidade Indisponível",
    "Motivo",
    "Valor Aplicado",
    "Valor bruto",
    "Valor líquido",
    "Valor Atualizado",
]

MOV_HEADER = [
    "Entrada/Saída",
    "Data",
    "Movimentação",
    "Produto",
    "Instituição",
    "Quantidade",
    "Preço unitário",
    "Valor da Operação",
]

INSTITUICAO = "INTER DTVM"

IPCA_2032 = [
    "Tesouro IPCA+ 2032", INSTITUICAO, "BRSTNCNTB7T1", "IPCA", "15/08/2032",
    15.48, 15.48, 0, "-", 44942.46, 45390.59, 45263.94, 45390.59,
]

PREFIXADO_2032 = [
    "Tesouro Prefixado 2032", INSTITUICAO, "BRSTNCLTN8J8", "prefixado", "01/01/2032",
    62.61, 62.61, 0, "-", 29995.82, 29802.98, 29766.19, 29802.98,
]

IPCA_BUY_ROW = [
    "Credito", "30/04/2026", "Compra", "Tesouro IPCA+ 2032",
    INSTITUICAO, 2.02, 2959.59, 5978.37,
]

PREFIXADO_BUY_ROW = [
    "Credito", "30/04/2026", "Compra", "Tesouro Prefixado 2032",
    INSTITUICAO, 12.46, 481.44, 5998.87,
]

NON_TD_ROW = [
    "Credito", "30/04/2026", "Direito de Subscrição",
    "BTLG12 - BTG PACTUAL LOGISTICA", INSTITUICAO, 209, "-", "-",
]

CDB_ROW = [
    "Credito", "29/04/2026", "COMPRA / VENDA", "CDB - CDB426DGCVL",
    INSTITUICAO, 10, 1000, 10000,
]


def _build_posicao(tmp_path: Path, rows: list[list]) -> str:
    path = tmp_path / "posicao.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Tesouro Direto"
    ws.append(POSICAO_HEADER)
    for row in rows:
        ws.append(row)
    ws.append([None] * len(POSICAO_HEADER))
    footer = [None] * len(POSICAO_HEADER)
    footer[POSICAO_HEADER.index("Valor Atualizado")] = "Total"
    ws.append(footer)
    wb.save(path)
    return str(path)


def _build_movimentacao(tmp_path: Path, rows: list[list]) -> str:
    path = tmp_path / "movimentacao.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentação"
    ws.append(MOV_HEADER)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return str(path)


def test_happy_path_parses_tesouro_positions(tmp_path):
    path = _build_posicao(tmp_path, [IPCA_2032, PREFIXADO_2032])

    positions = parse_tesouro_positions(path)

    assert len(positions) == 2
    ipca = positions[0]
    assert ipca.name == "Tesouro IPCA+ 2032"
    assert ipca.isin == "BRSTNCNTB7T1"
    assert ipca.indexer == "IPCA"
    assert ipca.maturity_date == date(2032, 8, 15)
    assert ipca.quantity == Decimal("15.48")
    assert ipca.current_value == Decimal("45390.59")
    # current_price is quantized to AssetMetaData's 10 decimal places.
    assert ipca.current_price == Decimal("2932.2086563307")


def test_position_blank_and_total_rows_are_skipped(tmp_path):
    path = _build_posicao(tmp_path, [IPCA_2032])

    positions = parse_tesouro_positions(path)

    assert len(positions) == 1


def test_position_missing_sheet_raises(tmp_path):
    path = tmp_path / "x.xlsx"
    wb = Workbook()
    wb.active.title = "Other"
    wb.save(path)

    with pytest.raises(B3ParserError):
        parse_tesouro_positions(str(path))


def test_position_missing_required_header_raises(tmp_path):
    bad_header = [h for h in POSICAO_HEADER if h != "Quantidade"]
    bad_row = [c for h, c in zip(POSICAO_HEADER, IPCA_2032) if h != "Quantidade"]
    path = tmp_path / "x.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Tesouro Direto"
    ws.append(bad_header)
    ws.append(bad_row)
    wb.save(path)

    with pytest.raises(B3ParserError, match="Quantidade"):
        parse_tesouro_positions(str(path))


def test_movements_filter_to_tesouro_compra_venda(tmp_path):
    path = _build_movimentacao(
        tmp_path,
        [NON_TD_ROW, CDB_ROW, IPCA_BUY_ROW, PREFIXADO_BUY_ROW],
    )

    movements = parse_tesouro_movements(path)

    assert len(movements) == 2
    assert movements[0].name == "Tesouro IPCA+ 2032"
    assert movements[0].action == B3FixedIncomeAction.BUY
    assert movements[0].operation_date == date(2026, 4, 30)
    assert movements[0].quantity == Decimal("2.02")
    assert movements[0].unit_price == Decimal("2959.59")
    assert movements[1].name == "Tesouro Prefixado 2032"


def test_movements_venda_maps_to_sell(tmp_path):
    sell_row = list(IPCA_BUY_ROW)
    sell_row[0] = "Debito"
    sell_row[2] = "Venda"
    path = _build_movimentacao(tmp_path, [sell_row])

    movements = parse_tesouro_movements(path)

    assert len(movements) == 1
    assert movements[0].action == B3FixedIncomeAction.SELL


def test_movements_flow_disagreement_raises(tmp_path):
    bad = list(IPCA_BUY_ROW)
    bad[0] = "Debito"
    path = _build_movimentacao(tmp_path, [bad])

    with pytest.raises(B3ParserError, match="disagrees"):
        parse_tesouro_movements(path)
