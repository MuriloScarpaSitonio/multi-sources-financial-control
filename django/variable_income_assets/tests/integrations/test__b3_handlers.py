from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from django.utils import timezone
from openpyxl import Workbook

from ...integrations.b3.handlers import (
    B3ImportError,
    import_b3_fixed_income_positions,
    import_b3_negociacoes,
)
from ...models import Asset, AssetMetaData, Transaction
from ...choices import AssetTypes, Currencies, LiquidityTypes, AssetObjectives
from ..conftest import AssetFactory, AssetMetaDataFactory

pytestmark = pytest.mark.django_db

POSICAO_HEADER = [
    "Produto", "Instituição", "Emissor", "Código", "Indexador", "Tipo de regime",
    "Data de Emissão", "Vencimento", "Quantidade", "Quantidade Disponível",
    "Quantidade Indisponível", "Motivo", "Contraparte",
    "Preço Atualizado MTM", "Valor Atualizado MTM",
    "Preço Atualizado CURVA", "Valor Atualizado CURVA",
    "Preço Atualizado FECHAMENTO", "Valor Atualizado FECHAMENTO",
]

TD_POSICAO_HEADER = [
    "Produto", "Instituição", "Código ISIN", "Indexador", "Vencimento",
    "Quantidade", "Quantidade Disponível", "Quantidade Indisponível", "Motivo",
    "Valor Aplicado", "Valor bruto", "Valor líquido", "Valor Atualizado",
]

ACOES_HEADER = [
    "Produto", "Instituição", "Conta", "Código de Negociação", "CNPJ da Empresa",
    "Código ISIN / Distribuição", "Tipo", "Escriturador", "Quantidade",
    "Quantidade Disponível", "Quantidade Indisponível", "Motivo",
    "Preço de Fechamento", "Valor Atualizado",
]

FII_HEADER = [
    "Produto", "Instituição", "Conta", "Código de Negociação", "CNPJ do Fundo",
    "Código ISIN / Distribuição", "Tipo", "Administrador", "Quantidade",
    "Quantidade Disponível", "Quantidade Indisponível", "Motivo",
    "Preço de Fechamento", "Valor Atualizado",
]

NEGOCIACAO_HEADER = [
    "Data do Negócio", "Tipo de Movimentação", "Mercado", "Prazo/Vencimento",
    "Instituição", "Código de Negociação", "Quantidade", "Preço", "Valor",
]

MOV_HEADER = [
    "Entrada/Saída", "Data", "Movimentação", "Produto", "Instituição",
    "Quantidade", "Preço unitário", "Valor da Operação",
]

POSICAO_FILENAME = "posicao-2026-04-29-12-00-00.xlsx"
WORKBOOK_DT = datetime(2026, 4, 29, 12, 0, 0)


def _cdb_position_row(*, code: str = "CDB426DGCVL", issuer: str = "BANCO BMG S/A"):
    return [
        f"CDB - {issuer}", "INTER DTVM", issuer, code, "PREFIXADO", "DEPOSITADO",
        "29/04/2026", "30/04/2029", "10", "10", "-", "-", "-",
        "-", "-", "1100", "11000", "-", "-",
    ]


def _cdb_movimentacao_row(*, code: str = "CDB426DGCVL"):
    return [
        "Credito", "29/04/2026", "COMPRA / VENDA", f"CDB - {code}",
        "INTER DTVM", 10, 1000, 10000,
    ]


def _build_posicao(
    tmp_path: Path,
    rows: list[list],
    td_rows: list[list] | None = None,
    acoes_rows: list[list] | None = None,
    fii_rows: list[list] | None = None,
) -> str:
    path = tmp_path / POSICAO_FILENAME
    wb = Workbook()
    ws = wb.active
    ws.title = "Renda Fixa"
    ws.append(POSICAO_HEADER)
    for row in rows:
        ws.append(row)
    td_sheet = wb.create_sheet("Tesouro Direto")
    td_sheet.append(TD_POSICAO_HEADER)
    for row in td_rows or []:
        td_sheet.append(row)
    acoes_sheet = wb.create_sheet("Acoes")
    acoes_sheet.append(ACOES_HEADER)
    for row in acoes_rows or []:
        acoes_sheet.append(row)
    fii_sheet = wb.create_sheet("Fundo de Investimento")
    fii_sheet.append(FII_HEADER)
    for row in fii_rows or []:
        fii_sheet.append(row)
    wb.save(path)
    return str(path)


def _build_negociacao(tmp_path: Path, rows: list[list]) -> str:
    path = tmp_path / "negociacao.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Negociação"
    ws.append(NEGOCIACAO_HEADER)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return str(path)


def _build_movimentacao(tmp_path: Path, rows: list[list]) -> str:
    path = tmp_path / "movimentacao-2026-04-29-12-00-00.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimentação"
    ws.append(MOV_HEADER)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return str(path)


@pytest.fixture
def fixed_br_asset(user):
    asset = AssetFactory(
        code="CDB426DGCVL",
        type=AssetTypes.fixed_br,
        currency=Currencies.real,
        objective=AssetObjectives.growth,
        user=user,
        liquidity_type=LiquidityTypes.at_maturity,
    )
    AssetMetaDataFactory(
        code="CDB426DGCVL",
        type=AssetTypes.fixed_br,
        currency=Currencies.real,
        current_price=Decimal("1000"),
        current_price_updated_at=timezone.make_aware(WORKBOOK_DT - timedelta(days=1)),
    )
    return asset


def test_existing_asset_stale_price_is_updated(tmp_path, user, fixed_br_asset):
    posicao_path = _build_posicao(tmp_path, [_cdb_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert len(report["actions"]) == 1
    assert report["actions"][0]["action"] == "price_updated"
    assert report["actions"][0]["new_price"] == "1100"
    metadata = AssetMetaData.objects.get(code="CDB426DGCVL")
    assert metadata.current_price == Decimal("1100")


def test_existing_asset_fresh_price_is_skipped(tmp_path, user, fixed_br_asset):
    metadata = AssetMetaData.objects.get(code="CDB426DGCVL")
    metadata.current_price_updated_at = timezone.make_aware(WORKBOOK_DT + timedelta(hours=1))
    metadata.save()
    posicao_path = _build_posicao(tmp_path, [_cdb_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["actions"][0]["action"] == "price_skipped"
    metadata.refresh_from_db()
    assert metadata.current_price == Decimal("1000")


def test_missing_asset_with_movimentacao_creates_asset_and_transaction(
    tmp_path, user, sync_assets_read_model
):
    posicao_path = _build_posicao(tmp_path, [_cdb_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [_cdb_movimentacao_row()])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["actions"][0]["action"] == "created"
    asset = Asset.objects.get(user=user, code="CDB426DGCVL")
    assert asset.type == AssetTypes.fixed_br
    assert asset.description == "CDB - BANCO BMG - venc 30/04/2029"
    assert asset.liquidity_type == LiquidityTypes.at_maturity
    transaction = Transaction.objects.get(asset=asset)
    assert transaction.price == Decimal("1000")
    assert transaction.quantity == Decimal("10")


def test_missing_asset_without_movimentacao_is_skipped_by_default(tmp_path, user):
    posicao_path = _build_posicao(tmp_path, [_cdb_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["actions"][0]["action"] == "skipped"
    assert "no movimentação" in report["actions"][0]["reason"]
    assert not Asset.objects.filter(user=user, code="CDB426DGCVL").exists()


def test_missing_asset_without_movimentacao_with_fallback_uses_posicao_price(
    tmp_path, user, sync_assets_read_model
):
    posicao_path = _build_posicao(tmp_path, [_cdb_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        use_posicao_price_when_missing_movement=True,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["actions"][0]["action"] == "created"
    transaction = Transaction.objects.get(asset__user=user, asset__code="CDB426DGCVL")
    assert transaction.price == Decimal("1100")
    assert transaction.quantity == Decimal("10")


def test_dry_run_rolls_back(tmp_path, user, sync_assets_read_model):
    posicao_path = _build_posicao(tmp_path, [_cdb_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [_cdb_movimentacao_row()])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=True,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["dry_run"] is True
    assert report["actions"][0]["action"] == "created"
    assert not Asset.objects.filter(user=user, code="CDB426DGCVL").exists()


def _td_position_row(*, name: str = "Tesouro IPCA+ 2032", isin: str = "BRSTNCNTB7T1"):
    return [
        name, "INTER DTVM", isin, "IPCA", "15/08/2032",
        15.48, 15.48, 0, "-", 44942.46, 45390.59, 45263.94, 45390.59,
    ]


def _td_movimentacao_row(*, name: str = "Tesouro IPCA+ 2032"):
    return [
        "Credito", "30/04/2026", "Compra", name,
        "INTER DTVM", 2.02, 2959.59, 5978.37,
    ]


@pytest.fixture
def td_asset(user):
    asset = AssetFactory(
        code="BRSTNCNTB7T1",
        type=AssetTypes.fixed_br,
        currency=Currencies.real,
        objective=AssetObjectives.growth,
        user=user,
        liquidity_type=LiquidityTypes.at_maturity,
    )
    AssetMetaDataFactory(
        code="BRSTNCNTB7T1",
        type=AssetTypes.fixed_br,
        currency=Currencies.real,
        current_price=Decimal("2900"),
        current_price_updated_at=timezone.make_aware(WORKBOOK_DT - timedelta(days=1)),
    )
    return asset


def test_existing_td_asset_with_new_movement_creates_transaction(
    tmp_path, user, td_asset, sync_assets_read_model
):
    posicao_path = _build_posicao(tmp_path, [], td_rows=[_td_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [_td_movimentacao_row()])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    actions_by_kind = {a["action"] for a in report["actions"]}
    assert "price_updated" in actions_by_kind
    assert "transaction_created" in actions_by_kind
    assert Transaction.objects.filter(
        asset=td_asset,
        operation_date=date(2026, 4, 30),
        quantity=Decimal("2.02"),
        price=Decimal("2959.59"),
    ).exists()


def test_existing_td_asset_with_already_imported_movement_is_not_duplicated(
    tmp_path, user, td_asset, sync_assets_read_model
):
    from ..conftest import TransactionFactory

    TransactionFactory(
        asset=td_asset,
        action="BUY",
        operation_date=date(2026, 4, 30),
        quantity=Decimal("2.02"),
        price=Decimal("2959.59"),
    )
    posicao_path = _build_posicao(tmp_path, [], td_rows=[_td_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [_td_movimentacao_row()])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    actions_by_kind = [a["action"] for a in report["actions"]]
    assert actions_by_kind.count("transaction_created") == 0
    assert Transaction.objects.filter(asset=td_asset).count() == 1


def test_existing_td_asset_stale_price_is_updated(tmp_path, user, td_asset):
    posicao_path = _build_posicao(tmp_path, [], td_rows=[_td_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert len(report["actions"]) == 1
    assert report["actions"][0]["action"] == "price_updated"
    assert report["actions"][0]["code"] == "BRSTNCNTB7T1"
    metadata = AssetMetaData.objects.get(code="BRSTNCNTB7T1")
    assert metadata.current_price == Decimal("2932.2086563308")


def test_missing_td_asset_with_movimentacao_creates_asset_and_transaction(
    tmp_path, user, sync_assets_read_model
):
    posicao_path = _build_posicao(tmp_path, [], td_rows=[_td_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [_td_movimentacao_row()])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["actions"][0]["action"] == "created"
    asset = Asset.objects.get(user=user, code="BRSTNCNTB7T1")
    assert asset.type == AssetTypes.fixed_br
    assert asset.description == "Tesouro IPCA+ 2032 - venc 15/08/2032"
    transaction = Transaction.objects.get(asset=asset)
    assert transaction.price == Decimal("2959.59")
    assert transaction.quantity == Decimal("2.02")


def test_missing_td_asset_without_movimentacao_is_skipped(tmp_path, user):
    posicao_path = _build_posicao(tmp_path, [], td_rows=[_td_position_row()])
    movimentacao_path = _build_movimentacao(tmp_path, [])

    report = import_b3_fixed_income_positions(
        user_id=user.id,
        dry_run=False,
        posicao_path=posicao_path,
        movimentacao_path=movimentacao_path,
    )

    assert report["actions"][0]["action"] == "skipped"
    assert "no movimentação" in report["actions"][0]["reason"]
    assert not Asset.objects.filter(user=user, code="BRSTNCNTB7T1").exists()


def _negotiation_row(*, code: str = "BBAS3"):
    return [
        "02/04/2026", "Compra", "Mercado à Vista", "-", "INTER DTVM",
        code, 100, 23.39, 2339,
    ]


def test_negociacao_creates_transaction_for_existing_asset(tmp_path, user):
    AssetFactory(
        code="BBAS3", type=AssetTypes.stock, currency=Currencies.real,
        objective=AssetObjectives.growth, user=user,
    )
    AssetMetaDataFactory(
        code="BBAS3", type=AssetTypes.stock, currency=Currencies.real,
        current_price=Decimal("21.71"), current_price_updated_at=timezone.now(),
    )
    from ...management.commands.sync_assets_cqrs import Command as Sync
    Sync().handle(user_ids=[user.id])
    negociacao_path = _build_negociacao(tmp_path, [_negotiation_row()])

    report = import_b3_negociacoes(
        user_id=user.id, dry_run=False, negociacao_path=negociacao_path
    )

    assert len(report["actions"]) == 1
    assert report["actions"][0]["action"] == "transaction_created"
    transaction = Transaction.objects.get(asset__user=user, asset__code="BBAS3")
    assert transaction.price == Decimal("23.39")
    assert transaction.quantity == Decimal("100")


def test_negociacao_skips_when_asset_not_in_db(tmp_path, user):
    negociacao_path = _build_negociacao(tmp_path, [_negotiation_row()])

    report = import_b3_negociacoes(
        user_id=user.id, dry_run=False, negociacao_path=negociacao_path
    )

    assert report["actions"][0]["action"] == "skipped"
    assert "asset not in DB" in report["actions"][0]["reason"]
    assert not Transaction.objects.filter(asset__user=user).exists()


def test_negociacao_dedupes_already_imported_transactions(tmp_path, user):
    from ..conftest import TransactionFactory

    asset = AssetFactory(
        code="BBAS3", type=AssetTypes.stock, currency=Currencies.real,
        objective=AssetObjectives.growth, user=user,
    )
    TransactionFactory(
        asset=asset, action="BUY",
        operation_date=date(2026, 4, 2),
        quantity=Decimal("100"), price=Decimal("23.39"),
    )
    negociacao_path = _build_negociacao(tmp_path, [_negotiation_row()])

    report = import_b3_negociacoes(
        user_id=user.id, dry_run=False, negociacao_path=negociacao_path
    )

    assert all(a["action"] != "transaction_created" for a in report["actions"])
    assert Transaction.objects.filter(asset=asset).count() == 1


def test_negociacao_creates_asset_when_missing_with_posicao(
    tmp_path, user, sync_assets_read_model
):
    acoes_row = [
        "BBAS3 - BCO BRASIL S.A.", "INTER DTVM", "4038379", "BBAS3",
        "00000000000191", "BRBBASACNOR3 - 337", "ON", "BANCO DO BRASIL S/A",
        971, 971, "-", "-", 21.71, 21080.41,
    ]
    posicao_path = _build_posicao(tmp_path, [], acoes_rows=[acoes_row])
    negociacao_path = _build_negociacao(tmp_path, [_negotiation_row(code="BBAS3")])

    report = import_b3_negociacoes(
        user_id=user.id, dry_run=False, create_missing_assets=True,
        negociacao_path=negociacao_path, posicao_path=posicao_path,
    )

    actions = [a["action"] for a in report["actions"]]
    assert actions == ["asset_created", "transaction_created"]
    asset = Asset.objects.get(user=user, code="BBAS3")
    assert asset.type == AssetTypes.stock
    assert Transaction.objects.filter(asset=asset).count() == 1


def test_negociacao_skips_when_missing_and_no_posicao(tmp_path, user):
    negociacao_path = _build_negociacao(tmp_path, [_negotiation_row()])

    report = import_b3_negociacoes(
        user_id=user.id, dry_run=False, negociacao_path=negociacao_path
    )

    assert report["actions"][0]["action"] == "skipped"
    assert "create_missing_assets" in report["actions"][0]["reason"]


def test_negociacao_dry_run_rolls_back(tmp_path, user):
    AssetFactory(
        code="BBAS3", type=AssetTypes.stock, currency=Currencies.real,
        objective=AssetObjectives.growth, user=user,
    )
    AssetMetaDataFactory(
        code="BBAS3", type=AssetTypes.stock, currency=Currencies.real,
        current_price=Decimal("21.71"), current_price_updated_at=timezone.now(),
    )
    from ...management.commands.sync_assets_cqrs import Command as Sync
    Sync().handle(user_ids=[user.id])
    negociacao_path = _build_negociacao(tmp_path, [_negotiation_row()])

    report = import_b3_negociacoes(
        user_id=user.id, dry_run=True, negociacao_path=negociacao_path
    )

    assert report["dry_run"] is True
    assert report["actions"][0]["action"] == "transaction_created"
    assert not Transaction.objects.filter(asset__user=user).exists()


def test_invalid_posicao_filename_raises(tmp_path, user):
    bad = tmp_path / "not-a-posicao.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Renda Fixa"
    ws.append(POSICAO_HEADER)
    ws.append(_cdb_position_row())
    wb.save(bad)
    movimentacao_path = _build_movimentacao(tmp_path, [])

    with pytest.raises(B3ImportError, match="filename"):
        import_b3_fixed_income_positions(
            user_id=user.id,
            posicao_path=str(bad),
            movimentacao_path=movimentacao_path,
        )
